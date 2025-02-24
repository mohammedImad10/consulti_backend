import asyncio
import os
import json
from platform import processor
import re
import hashlib
from time import sleep
import numpy as np
import pandas as pd
from pypdf import PdfReader
from dotenv import load_dotenv
import chromadb
from chromadb.utils.embedding_functions import OpenAIEmbeddingFunction
from openpyxl import load_workbook
from diskcache import Cache
from typing import List, Optional
import openai
from openai import OpenAI
from pydantic_ai import Agent, RunContext
from pydantic_ai.models.openai import OpenAIModel
from flask_socketio import SocketIO, emit
import threading

# Global DEBUG flag: set to True to show debug logs
DEBUG = True
user_response = None
waiting_for_response = False
response_event = asyncio.Event()
global_socketio = None
# =================== LOAD ENVIRONMENT VARIABLES ===================
load_dotenv()
openai_key = os.getenv("OPENAI_API_KEY")
if not openai_key:
    raise ValueError("Please set your OPENAI_API_KEY in the environment variables.")

# =================== PERSISTENT CACHE CLASS ===================
class PersistentCache:
    """A wrapper around diskcache for persistent caching with TTL."""
    def __init__(self, cache_dir: str = "./cache", ttl: int = 3600):
        self.cache = Cache(cache_dir)
        self.ttl = ttl

    def get(self, key: str):
        return self.cache.get(key)

    def set(self, key: str, value, ttl: Optional[int] = None):
        self.cache.set(key, value, expire=ttl or self.ttl)

    def clear(self):
        self.cache.clear()

    def __del__(self):
        self.cache.close()

# =================== AGENTIC RAG CLASS ===================
class AgenticRAG:
    def __init__(self, collection_name: str, is_client: bool):
        self.openai_key = openai_key
        self.client = OpenAI(api_key=self.openai_key)
        self.llm_model_name = os.getenv('LLM_MODEL', 'gpt-4o-mini')
        self.model = OpenAIModel(self.llm_model_name)
        # For global docs (yellow) use is_client=False; for customer docs (pink) use is_client=True.
        self.db_path = "./clients_db_storage" if is_client else "./chromadb_storage"
        self.chroma_client = chromadb.PersistentClient(path=self.db_path)
        self.collection = self.chroma_client.get_or_create_collection(
            name=collection_name,
            embedding_function=OpenAIEmbeddingFunction(api_key=openai_key)
        )
        self.system_prompt = (
            "You are an AI assistant that helps process documents by extracting content, generating embeddings, "
            "and retrieving relevant information. You may retrieve, generate, or refine responses iteratively based on the query."
        )
        self.knowledge_agent = Agent(self.model, system_prompt=self.system_prompt, retries=2)
        self.cache = PersistentCache(cache_dir="./cache", ttl=3600)
        self.response_cache = PersistentCache(cache_dir="./response_cache", ttl=3600)
        # Ingestion metadata to avoid reprocessing unchanged files.
        self.metadata_path = f"ingestion_metadata_{collection_name}.json"
        self.ingestion_metadata = self.load_ingestion_metadata()

    def load_ingestion_metadata(self) -> dict:
        if os.path.exists(self.metadata_path):
            try:
                with open(self.metadata_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception as e:
                print(f"[ERROR] Loading metadata failed: {e}")
        return {}

    def save_ingestion_metadata(self):
        try:
            with open(self.metadata_path, "w", encoding="utf-8") as f:
                json.dump(self.ingestion_metadata, f)
        except Exception as e:
            print(f"[ERROR] Saving metadata failed: {e}")

    # ------------------ DOCUMENT INGESTION METHODS ------------------
    async def extract_text_from_pdf(self, pdf_path: str) -> List[str]:
        try:
            reader = PdfReader(pdf_path)
            text_chunks = []
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    chunks = [chunk.strip() for chunk in page_text.split("\n\n") if chunk.strip()]
                    text_chunks.extend(chunks)
            print(f"[INFO] Extracted {len(text_chunks)} chunks from PDF: {pdf_path}")
            return text_chunks
        except Exception as e:
            print(f"[ERROR] Failed to extract PDF {pdf_path}: {e}")
            return []

    async def extract_data_from_excel(self, excel_path: str) -> List[str]:
        try:
            df = pd.read_excel(excel_path)
            chunks = df.astype(str).apply(lambda row: " | ".join(row), axis=1).tolist()
            print(f"[INFO] Extracted {len(chunks)} chunks from Excel: {excel_path}")
            return chunks
        except Exception as e:
            print(f"[ERROR] Failed to extract Excel {excel_path}: {e}")
            return []

    async def ingest_pdf(self, pdf_path: str) -> None:
        mod_time = os.path.getmtime(pdf_path)
        if pdf_path in self.ingestion_metadata and self.ingestion_metadata[pdf_path] >= mod_time:
            print(f"[INFO] PDF unchanged: {pdf_path}")
            return
        text_chunks = await self.extract_text_from_pdf(pdf_path)
        if not text_chunks:
            print(f"[WARN] No chunks found in PDF: {pdf_path}")
            return
        for i, chunk in enumerate(text_chunks):
            self.collection.add(ids=[f"pdf_{os.path.basename(pdf_path)}_{i}"], documents=[chunk])
        self.ingestion_metadata[pdf_path] = mod_time
        self.save_ingestion_metadata()
        self.cache.clear()
        self.response_cache.clear()
        print(f"[INFO] Ingested PDF: {pdf_path} with {len(text_chunks)} chunks.")

    async def ingest_excel(self, excel_path: str) -> None:
        mod_time = os.path.getmtime(excel_path)
        if excel_path in self.ingestion_metadata and self.ingestion_metadata[excel_path] >= mod_time:
            print(f"[INFO] Excel unchanged: {excel_path}")
            return
        data_chunks = await self.extract_data_from_excel(excel_path)
        if not data_chunks:
            print(f"[WARN] No data extracted from Excel: {excel_path}")
            return
        for i, chunk in enumerate(data_chunks):
            self.collection.add(ids=[f"excel_{os.path.basename(excel_path)}_{i}"], documents=[chunk])
        self.ingestion_metadata[excel_path] = mod_time
        self.save_ingestion_metadata()
        self.cache.clear()
        self.response_cache.clear()
        print(f"[INFO] Ingested Excel: {excel_path} with {len(data_chunks)} chunks.")

    async def ingest_txt(self, txt_path: str) -> None:
        mod_time = os.path.getmtime(txt_path)
        if txt_path in self.ingestion_metadata and self.ingestion_metadata[txt_path] >= mod_time:
            print(f"[INFO] Text file unchanged: {txt_path}")
            return
        try:
            with open(txt_path, "r", encoding="utf-8") as f:
                text = f.read()
            doc_id = f"txt_{os.path.basename(txt_path)}_0"
            # Remove previous document if exists
            self.collection.delete(ids=[doc_id])
            self.collection.add(ids=[doc_id], documents=[text])
            self.ingestion_metadata[txt_path] = mod_time
            self.save_ingestion_metadata()
            self.cache.clear()
            self.response_cache.clear()
            print(f"[INFO] Ingested text file: {txt_path}")
        except Exception as e:
            print(f"[ERROR] Failed to ingest text file {txt_path}: {e}")

    async def ingest_directory(self, directory: str) -> None:
        print(f"[INFO] Ingesting from directory: {directory}")
        for filename in os.listdir(directory):
            file_path = os.path.join(directory, filename)
            if os.path.isfile(file_path):
                ext = filename.lower().split(".")[-1]
                if ext == "pdf":
                    await self.ingest_pdf(file_path)
                elif ext in ["xlsx", "xls"]:
                    await self.ingest_excel(file_path)
                elif ext == "txt":
                    await self.ingest_txt(file_path)
                else:
                    print(f"[WARN] Skipped unsupported file: {filename}")
        print(f"[INFO] Finished ingesting from directory: {directory}")

    # ------------------ QUERY & RESPONSE METHODS ------------------
    async def query_data(self, context: RunContext, user_query: str, n_results: int = 5) -> list:
        try:
            query_hash = hashlib.md5(user_query.encode()).hexdigest()
            cached = self.cache.get(query_hash)
            if cached:
                if DEBUG:
                    print(f"[DEBUG] Using cached chunks for query: {user_query}")
                return cached
            results = self.collection.query(query_texts=[user_query], n_results=n_results)
            retrieved = results["documents"][0] if results.get("documents") else []
            self.cache.set(query_hash, retrieved)
            if DEBUG:
                print(f"[DEBUG] Found {len(retrieved)} chunks for query: {user_query}")
            return retrieved
        except Exception as e:
            print(f"[ERROR] Error retrieving data for '{user_query}': {e}")
            return []

    async def generate_response(self, query: str, context_chunks: Optional[list], searching_web: bool = False) -> str:
        """
        If searching_web=True, the prompt includes a statement that we are 'searching the web' and should provide references.
        """
        query_hash = hashlib.md5(query.encode()).hexdigest()
        # cached_response = self.response_cache.get(query_hash)
        # if cached_response:
        #     print(f"[DEBUG] Using cached response")
        #     return cached_response
        context_text = "\n\n".join(context_chunks) if context_chunks else "No context provided."
        web_note = ""
        if not context_chunks and searching_web:
            web_note = (
                "No local data found. Searching the web for typical values and references. "
                "Include references in your final answer."
            )
        prompt = (
            "You are an assistant for question-answering tasks. Use the following context to answer the question. "
            "If the context contains multiple numerical values for the same attribute (e.g., area values), sum them and provide the total. "
            "Then, on a new line, write FINAL_ANSWER: followed by the exact final value (or 'N/A').\n\n"
            "if you didnt find anything in the knowldege base seatch the web"
            "Include references in your final answer."
            f"{web_note}\n"
            f"Context:\n{context_text}\n\nQuestion:\n{query}"
        )
        if DEBUG:
            print(f"[DEBUG] Prompt length: {len(prompt)} chars")
        response = self.client.chat.completions.create(
            model=self.llm_model_name,
            messages=[{"role": "system", "content": prompt}, {"role": "user", "content": query}],
            temperature=0.3
        )
        answer = response.choices[0].message.content
        self.response_cache.set(hashlib.md5(query.encode()).hexdigest(), answer)
        return answer

    async def agentic_rag(self, query: str, is_client: bool, unique_id: Optional[str] = None, max_iterations: int = 1) -> str:
        """
        1) We query local knowledge base. If empty & pink cell => we attempt "web search" approach.
        2) Return the final answer or 'N/A' if not found.
        """
        combined_query = f"{unique_id}:{query}" if unique_id else query
        run_context = RunContext(deps={}, model=self.model, usage={}, prompt=self.system_prompt)

        # 1) Query local knowledge base
        retrieved_chunks = await self.query_data(run_context, query, n_results=5)

        # 2) If no local data & is pink => simulate searching the web
        searching_web = (not is_client)

        # 3) Generate the final response
        response = await self.generate_response(combined_query, retrieved_chunks, searching_web=searching_web)
        if DEBUG:
            print(f"[DEBUG] agentic_rag final response:\n{response}")

        return response

# =================== LLM CLIENT CLASS ===================
class LLMClient:
    def __init__(self, api_key: str, system_prompt: str, model: str = "gpt-4o-mini", temperature: float = 0.3):
        self.api_key = api_key
        self.system_prompt = system_prompt
        self.model = model
        self.temperature = temperature

    def query(self, prompt: str) -> str:
        try:
            client = openai.OpenAI(api_key=self.api_key)
            response = client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": self.system_prompt},
                    {"role": "user", "content": prompt}
                ],
                temperature=self.temperature
            )
            return response.choices[0].message.content.strip()
        except Exception as e:
            print(f"[ERROR] LLM Error: {e}")
            return None

# =================== CHAT INTERFACE ===================
def chat_interface():
    # Define directories for knowledge ingestion.
    global_docs_dir = "global_docs"          # Global (book) knowledge
    customer_docs_dir = "customer_data_docs"   # Customer-specific documents

    # Create and ingest global and customer knowledge bases.
    print("[INFO] Ingesting global docs...")
    # socketio.emit('info', {'message': f'[INFO] Ingesting global docs...'})
    agentic_rag_global = AgenticRAG(collection_name="global_agent_knowledge", is_client=False)
    asyncio.run(agentic_rag_global.ingest_directory(global_docs_dir))
    print("[INFO] Finished ingesting global docs.")
    # socketio.emit('info', {'message': f'[INFO] finished ingestion global docs'})

    print("[INFO] Ingesting customer docs...")
    agentic_rag_customer = AgenticRAG(collection_name="customer_data_knowledge", is_client=True)
    asyncio.run(agentic_rag_customer.ingest_directory(customer_docs_dir))
    print("[INFO] Finished ingesting customer docs.")

# =================== EXCEL TEMPLATE PROCESSOR ===================
class ExcelTemplateProcessor:
    YELLOW_RGB = 'FFFFFF00'  # For yellow cells
    PINK_RGB = 'FFFF7DFF'    # For pink cells

    def __init__(self, file_path: str, llm_client: LLMClient,
                 agentic_rag_yellow: AgenticRAG, agentic_rag_pink: AgenticRAG):
        self.file_path = file_path
        self.llm_client = llm_client
        self.agentic_rag_yellow = agentic_rag_yellow
        self.agentic_rag_pink = agentic_rag_pink
        self.workbook, self.sheets_data = self._load_workbook()
        self.system_prompt = (
            "You are a financial analyst assistant. Your tasks include extracting precise values from provided text "
            "and benchmarking using available data. Return your answers as valid JSON with no additional commentary."
        )
        
    def _load_workbook(self):
        wb = load_workbook(self.file_path)
        sheets_data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_cells = {'yellow': [], 'pink': []}
            for row in ws.iter_rows():
                for cell in row:
                    fill = cell.fill
                    if fill and hasattr(fill, 'fgColor') and fill.fgColor and fill.fgColor.rgb:
                        rgb = fill.fgColor.rgb
                        if rgb == self.YELLOW_RGB:
                            sheet_cells['yellow'].append(cell)
                        elif rgb == self.PINK_RGB:
                            sheet_cells['pink'].append(cell)
            sheets_data[sheet_name] = sheet_cells
        return wb, sheets_data

    def get_cell_context(self, cell, header_search_range=3, unit_search_range=5):
        ws = cell.parent
        header = None
        for offset in range(1, header_search_range + 1):
            col_index = cell.column - offset
            if col_index < 1:
                break
            candidate = ws.cell(row=cell.row, column=cell.column - offset).value
            if candidate and isinstance(candidate, str) and candidate.strip():
                header = candidate.strip()
                break
        if not header and cell.row > 1:
            candidate = ws.cell(row=cell.row - 1, column=cell.column).value
            if candidate and isinstance(candidate, str) and candidate.strip():
                header = candidate.strip()
        unit = None
        for offset in range(1, unit_search_range + 1):
            row_index = cell.row - offset
            if row_index < 1:
                break
            candidate = ws.cell(row=row_index, column=cell.column).value
            if candidate and isinstance(candidate, str):
                if any(k in candidate.lower() for k in ["meter", "m", "%", "dollar", "$", "€", "£"]):
                    unit = candidate.strip()
                    break
        return header, unit

    def llm_query(self, prompt, model="gpt-4o-mini"):
        try:
            client = openai.OpenAI(api_key=openai_key)
            response = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": self.system_prompt},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.3
            )
            return response.choices[0].message.content.strip()
        except Exception as e:
            print(f"[ERROR] LLM Error: {e}")
            return None

    def get_question_for_cell(self, cell, header, unit):
        context_info = f"Field Description: {header or 'N/A'}"
        if unit:
            context_info += f"; Unit: {unit}"
        prompt = (
            f"Generate one concise, professional question to clarify the required value for the field with the following context:\n\n"
            f"{context_info}\n\nDo not reference any cell coordinates."
            "Do not reference any cell coordinates."
            "if you dont find any answer just return the value is N/A "
        )
        question = self.llm_query(prompt)
        return question

    def convert_value(self, val: str):
        val = val.strip()
        if "%" in val:
            return val
        try:
            num = float(val)
            return int(num) if num.is_integer() else num
        except Exception:
            return val

    def parse_final_answer(self, response: str) -> Optional[str]:
        match = re.search(r'(?m)^FINAL_ANSWER:\s*(.*)$', response)
        if match:
            return match.group(1).strip()
        return None

    async def fill_yellow_cells_from_docs(self, socketio, max_fields=6):
        processed_fields = 0
        for sheet_name, data in self.sheets_data.items():
            for cell in data['yellow']:
                if processed_fields >= max_fields:
                    print("[INFO] Processed maximum yellow cells.")
                    return
                if cell.value and isinstance(cell.value, str) and cell.value.strip().lower() != "n/a":
                    continue
                processed_fields += 1
                header, unit = self.get_cell_context(cell)
                if not header:
                    header = "Field with no header"
                field_id = f"{sheet_name}_{cell.coordinate}"
                print(f"[INFO] Yellow cell {field_id}, Header='{header}'")
                question = self.get_question_for_cell(cell, header, unit)

                # Single or multi-step approach (try 3 times).
                response = self.agentic_rag_yellow.agentic_rag(query=question, is_client=True, unique_id=field_id, max_iterations=3)
                
                final_answer = self.parse_final_answer(response)
                if not final_answer or final_answer.lower() in ["n/a", "i don't know", "no context provided", "no information provided"]:
                    print(f"[WARN] No valid FINAL_ANSWER for {header} field {field_id}. Prompt user.\n")
                    # Emit the question to the client
                    print( f"[INFO] Emitting question: {question}")
                    socketio.emit('message', question)
                    
                    # Wait for the user's response
                    global response_event, user_response
                    response_event.clear()  # Clear the event before waiting
                    print("Waiting for user response...")
                    await response_event.wait()  # Wait for the event to be set
                    print("User response received.")

                    # Process the user's response
                    user_val = user_response
                    user_val = user_response
                    cell.value = self.convert_value(user_val) if user_val.strip() else "N/A"
                else:
                    cell.value = self.convert_value(final_answer)
                print(f"[INFO] Yellow cell {field_id} set to '{cell.value}'.\n\n")

    
    async def fill_pink_cells_from_docs(self, socketio, max_fields=6):
        processed_fields = 0
        for sheet_name, data in self.sheets_data.items():
            for cell in data['pink']:
                if processed_fields >= max_fields:
                    print("[INFO] Processed maximum pink cells.")
                    return
                if cell.value and isinstance(cell.value, str) and cell.value.strip().lower() != "n/a":
                    continue
                processed_fields += 1
                header, unit = self.get_cell_context(cell)
                if not header:
                    header = "Unknown metric"
                field_id = f"{sheet_name}_{cell.coordinate}"
                print(f"[INFO] Pink cell {field_id}, Header='{header}'")
                question = self.get_question_for_cell(cell, header, unit)

                # Pink cells: only 1 iteration to keep the final answer short.
                response = await self.agentic_rag_pink.agentic_rag(query=question, is_client=False, unique_id=field_id, max_iterations=1)
                print(f"[INFO] Agentic RAG response: {response}")
                final_answer = self.parse_final_answer(response)
                if not final_answer or final_answer.lower() in ["n/a", "i don't know", "no context provided", "no information provided"]:
                    print(f"[WARN] No valid FINAL_ANSWER for {header} field {field_id}. Prompt user.\n")
                    # Emit the question to the client
                    print( f"[INFO] Emitting question: {question}")
                    socketio.emit('message', question)
                    
                    # Wait for the user's response
                    global response_event, user_response
                    response_event.clear()  # Clear the event before waiting
                    print("Waiting for user response...")
                    await response_event.wait()  # Wait for the event to be set
                    print("User response received.")

                    # Process the user's response
                    user_val = user_response
                    print(f"[INFO] User response in cell fun: {user_val}")
                    cell.value = self.convert_value(user_val) if user_val.strip() else "N/A"
                else:
                    cell.value = self.convert_value(final_answer)
                print(f"[INFO] Pink cell {field_id} set to '{cell.value}'.\n\n")
    
    def save(self, output_file: str):
        self.workbook.save(output_file)
        print(f"[INFO] Workbook saved to {output_file}")

        
def set_user_response(response, loop):
    """Set the user's response and notify the process to continue."""
    global response_event, user_response
    user_response = response
    print(f"User response set: {user_response}")

    # Ensure the event is set in the correct event loop
    
    loop.call_soon_threadsafe(response_event.set)


# =================== CHAT INTERFACE ===================
def chat(agentic_rag_global, agentic_rag_customer):
    print("[INFO] Starting chat interface. Type 'exit' to quit.")
    while True:
        query = input("Enter your question: ").strip()
        if query.lower() == "exit":
            break
        kb_choice = input("Which knowledge base to query? (g for global, c for customer): ").strip().lower()
        if kb_choice in ["g", "global"]:
            is_client = False
            agentic_rag = agentic_rag_global
        elif kb_choice in ["c", "customer"]:
            is_client = True
            agentic_rag = agentic_rag_customer
        else:
            print("[WARN] Invalid choice; defaulting to global.")
            is_client = False
            agentic_rag = agentic_rag_global

        response = asyncio.run(agentic_rag.agentic_rag(query, is_client=is_client, unique_id=query, max_iterations=1))
        print("Bot response:", response)
        print("-" * 80)

# =================== MAIN FUNCTION ===================
async def main(socketio):
    global global_socketio
    global_socketio = socketio
    global_docs_dir = "global_docs"          # Global (book) knowledge
    customer_docs_dir = "customer_data_docs"   # Customer-specific documents

    print("[INFO] Ingesting global docs...")
    socketio.emit('message', f'[INFO] Ingesting global docs...')
    agentic_rag_global = AgenticRAG(collection_name="global_agent_knowledge", is_client=False)
    # asyncio.run(agentic_rag_global.ingest_directory(global_docs_dir))
    await agentic_rag_global.ingest_directory(global_docs_dir)
    print("[INFO] Finished ingesting global docs.")

    print("[INFO] Ingesting customer docs...")
    agentic_rag_customer = AgenticRAG(collection_name="customer_data_knowledge", is_client=True)
    # asyncio.run(agentic_rag_customer.ingest_directory(customer_docs_dir))
    await agentic_rag_customer.ingest_directory(customer_docs_dir)
    print("[INFO] Finished ingesting customer docs.")
    socketio.emit('message', {'message': f'[INFO] Finished ingesting customer docs'})


    system_prompt = (
        "You are a financial analyst assistant. You only return the final answer or 'N/A'. "
        "If searching the web, include references. No extra commentary is needed."
    )
    llm_client = LLMClient(api_key=openai_key, system_prompt=system_prompt, model="gpt-4o-mini", temperature=0.3)

    processor = ExcelTemplateProcessor("financial_sheet.xlsx", llm_client, agentic_rag_customer, agentic_rag_global)
    await processor.fill_yellow_cells_from_docs(max_fields=6)
    await processor.fill_pink_cells_from_docs(socketio= socketio,max_fields=60)
    processor.save("processed_sheet.xlsx")
    
    # chat_interface()

# if __name__ == "__main__":
#     main()
